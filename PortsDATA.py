import pyvista as pv
import numpy as np
import Surface_data as FuD
import NBI_Ports_data_input as Cout
import time
import openpyxl
import os
from scipy.spatial.transform import Rotation as R

def create_surface(R_x, R_y, R_z):
    points, faces = [], []
    for i in range(len(R_x) - 1):
        current_contour = np.column_stack((R_x[i], R_y[i], R_z[i]))
        next_contour = np.column_stack((R_x[i+1], R_y[i+1], R_z[i+1]))
        for j in range(len(current_contour)):
            p1, p2 = current_contour[j], current_contour[(j+1) % len(current_contour)]
            n1, n2 = next_contour[j], next_contour[(j+1) % len(current_contour)]
            points.extend([p1, p2, n2, n1])
            faces.append([4, len(points)-4, len(points)-3, len(points)-2, len(points)-1])
    return pv.PolyData(np.array(points), faces=np.hstack(faces))


def get_column_data(sheet, col_index):
    # Получаем данные из указанного столбца (индекс начинается с 1)
    column_data = [cell[0].value for cell in sheet.iter_rows(min_row=4, min_col=col_index, max_col=col_index, values_only=False)]
    return column_data

def delete(file_to_delete):

 if os.path.exists(file_to_delete):
    os.remove(file_to_delete)
    print(f"File {file_to_delete} deleted.")
 else:
    print(f"File {file_to_delete} not found.")  


def Module(output_sheet):
 for row in output_sheet.iter_rows(min_row=4, max_row=output_sheet.max_row, min_col=3, max_col=3):
    cell_value = row[0].value
    row_number = row[0].row
    #print(row_number)


    submodule_value = output_sheet[row_number][3].value
    if submodule_value == "rotation":
        submodule_value = 1
    #print(submodule_value)
    k = int(cell_value)*10+int(submodule_value)
    if k == 10 or k == 20 or k ==30 or k ==40 or k==50:
        l = 1
    else: 
     l = 0


    
    new_values = new_coordinates_for_Moduls(row_number, output_sheet, cell_value, l)
    for i in range(len(new_values)):
        output_sheet.cell(row=row_number, column=(i+5)).value = new_values[i]    
 return output_file_path


def create_new_file(input_file_path, output_file_path):
    

 input_workbook = openpyxl.load_workbook(input_file_path)
 input_sheet = input_workbook.active

 output_workbook = openpyxl.Workbook()
 output_sheet = output_workbook.active

 for row in input_sheet.iter_rows():
    output_sheet.append([cell.value for cell in row])
 return output_sheet, output_workbook


def translate_coord(X, Y, Z):

    R = np.sqrt(X**2 + Y**2)
    if X==0:
        Phi_0 = np.pi/2
    else:   
        Phi_0 = np.arctan(Y/X)
    
    Phi = np.degrees(Phi_0)
    Z_r = Z
    return R, Phi, Z_r

def back_translate_coord(R, Phi, Z):
    angle = np.radians(Phi)
    x = R * np.cos(angle)
    y = R * np.sin(angle)
    z = Z
    return x, y, z 
    

def new_coordinates_for_Moduls(row_number, output_sheet, cell_value, l):

    X_1, Y_1, Z_1 = output_sheet[row_number][4].value, output_sheet[row_number][5].value, output_sheet[row_number][6].value
    X_2, Y_2, Z_2 = output_sheet[row_number][7].value, output_sheet[row_number][8].value, output_sheet[row_number][9].value
    X_3, Y_3, Z_3 = output_sheet[row_number][10].value, output_sheet[row_number][11].value, output_sheet[row_number][12].value
    X_4, Y_4, Z_4 = output_sheet[row_number][13].value, output_sheet[row_number][14].value, output_sheet[row_number][15].value

    if l == 1:
       Y_1 = -Y_1
       Z_1 = -Z_1
       
       Y_2 = -Y_2
       Z_2 = -Z_2

       Y_3 = -Y_3
       Z_3 = -Z_3

       Y_4 = -Y_4
       Z_4 = -Z_4
    
    
    R_1, Phi_1, Z_1 =  translate_coord(X_1, Y_1, Z_1)
    R_2, Phi_2, Z_2 =  translate_coord(X_2, Y_2, Z_2)
    R_3, Phi_3, Z_3 =  translate_coord(X_3, Y_3, Z_3)
    R_4, Phi_4, Z_4 =  translate_coord(X_4, Y_4, Z_4)
    
    Phi_1 =   Phi_1  + 72* (cell_value -1) 
    Phi_2 =   Phi_2  + 72* (cell_value -1) 
    Phi_3 =   Phi_3  + 72* (cell_value -1) 
    Phi_4 =   Phi_4  + 72* (cell_value -1) 
    
    
    X_1, Y_1, Z_1 = back_translate_coord(R_1, Phi_1, Z_1)
    X_2, Y_2, Z_2 = back_translate_coord(R_2, Phi_2, Z_2)
    X_3, Y_3, Z_3 = back_translate_coord(R_3, Phi_3, Z_3)
    X_4, Y_4, Z_4 = back_translate_coord(R_4, Phi_4, Z_4)

    
    output_sheet[row_number][4].value = X_1
    output_sheet[row_number][5].value = Y_1
    output_sheet[row_number][6].value = Z_1
            
    output_sheet[row_number][7].value = X_2
    output_sheet[row_number][8].value = Y_2
    output_sheet[row_number][9].value = Z_2
            
    output_sheet[row_number][10].value = X_3
    output_sheet[row_number][11].value = Y_3
    output_sheet[row_number][12].value = Z_3
            
    output_sheet[row_number][13].value = X_4
    output_sheet[row_number][14].value = Y_4
    output_sheet[row_number][15].value = Z_4

    return [output_sheet[row_number][4].value, output_sheet[row_number][5].value, output_sheet[row_number][6].value,
            output_sheet[row_number][7].value, output_sheet[row_number][8].value, output_sheet[row_number][9].value,
            output_sheet[row_number][10].value, output_sheet[row_number][11].value, output_sheet[row_number][12].value,
            output_sheet[row_number][13].value, output_sheet[row_number][14].value, output_sheet[row_number][15].value]


def submodule(output_sheet):
    for row in output_sheet.iter_rows(min_row=4, max_row=output_sheet.max_row, min_col=3, max_col=3):
       module_value = row[0].value
       row_number = row[0].row
       submodule_value = output_sheet[row_number][3].value

       if submodule_value == "rotation":
        submodule_value = 1
       if submodule_value == 0:
           new_values2 = new_coordinates_for_SUBModuls(row_number, output_sheet, module_value, submodule_value)
           for i in range(len(new_values2)):
               output_sheet.cell(row=row_number, column=(i+5)).value = new_values2[i]
    return output_file_path 

def new_coordinates_for_SUBModuls(row_number, output_sheet, module_value, submodule_value):
    X_1 = output_sheet[row_number][4].value
    Y_1 = output_sheet[row_number][5].value
    Z_1 = output_sheet[row_number][6].value
    
    X_2 = output_sheet[row_number][7].value
    Y_2 = output_sheet[row_number][8].value
    Z_2 = output_sheet[row_number][9].value
    
    X_3 = output_sheet[row_number][10].value
    Y_3 = output_sheet[row_number][11].value
    Z_3 = output_sheet[row_number][12].value
    
    X_4 = output_sheet[row_number][13].value
    Y_4 = output_sheet[row_number][14].value
    Z_4 = output_sheet[row_number][15].value
    
    
    X_Zero = 0
    Y_Zero = 0 
    Z_Zero = 0
    
    
    R_two = 1
    Phi_two = 36 + 72 * (module_value-1)
    Z_two = 0
    
    X_two, Y_two, Z_two =  back_translate_coord(R_two,  Phi_two, Z_two)
    
    

    axis_point1 = np.array([X_Zero, Y_Zero, Z_Zero])
    axis_point2 = np.array([X_two, Y_two, Z_two])
    
    
    point1 = np.array([X_1, Y_1, Z_1])
    point2 = np.array([X_2, Y_2, Z_2])
    point3 = np.array([X_3, Y_3, Z_3])
    point4 = np.array([X_4, Y_4, Z_4])
    
    rotated_point1 = rotate_point_around_arbitrary_axis(point1, axis_point1, axis_point2)
    rotated_point2 = rotate_point_around_arbitrary_axis(point2, axis_point1, axis_point2)
    rotated_point3 = rotate_point_around_arbitrary_axis(point3, axis_point1, axis_point2)
    rotated_point4 = rotate_point_around_arbitrary_axis(point4, axis_point1, axis_point2)
    
    output_sheet[row_number][4].value = rotated_point1[0]
    output_sheet[row_number][5].value = rotated_point1[1]
    output_sheet[row_number][6].value = rotated_point1[2]
            
    output_sheet[row_number][7].value = rotated_point2[0]
    output_sheet[row_number][8].value = rotated_point2[1]
    output_sheet[row_number][9].value = rotated_point2[2]
            
    output_sheet[row_number][10].value = rotated_point3[0]
    output_sheet[row_number][11].value = rotated_point3[1]
    output_sheet[row_number][12].value = rotated_point3[2]
            
    output_sheet[row_number][13].value = rotated_point4[0]
    output_sheet[row_number][14].value = rotated_point4[1]
    output_sheet[row_number][15].value = rotated_point4[2]

    return [output_sheet[row_number][4].value, output_sheet[row_number][5].value, output_sheet[row_number][6].value,
            output_sheet[row_number][7].value, output_sheet[row_number][8].value, output_sheet[row_number][9].value,
            output_sheet[row_number][10].value, output_sheet[row_number][11].value, output_sheet[row_number][12].value,
            output_sheet[row_number][13].value, output_sheet[row_number][14].value, output_sheet[row_number][15].value]

def rotate_point_around_arbitrary_axis(point, axis_point1, axis_point2):
    angle_deg = 180
    angle_rad = np.radians(angle_deg)

    point = np.array(point)
    axis = np.array(axis_point2) - np.array(axis_point1)
    

    axis_length = np.linalg.norm(axis)
    axis_normalized = axis / axis_length
    

    rotation = R.from_rotvec(axis_normalized * angle_rad)
    

    rotated_point = rotation.apply(point - np.array(axis_point1)) + np.array(axis_point1)
    
    return np.array(rotated_point)

def name(output_sheet):
        P_name = get_column_data(output_sheet, 2)
        P_module = get_column_data(output_sheet, 3)
        P_submodule = get_column_data(output_sheet, 4)

        P_name = np.array(P_name)
        P_module = [int(value) for value in P_module]

        # Remove the first character from each port name
        modified_P_name = [name[2:] for name in P_name]  # Keep only the characters after the first one

        # Convert P_submodule to the desired format
        modified_P_submodule = [0 if submodule == 0 else 1 for submodule in P_submodule]

        # Create the new array with the desired format
        combined_array = [f"{module}_{submodule}_{name}" for module, submodule, name in zip(P_module, modified_P_submodule, modified_P_name)]

        return(combined_array)

def add_labels(plotter, points, labels, text_color='white', point_color='blue'):
    plotter.add_point_labels(pv.PolyData(points.T), labels, point_size=10, font_size=12, text_color=text_color, point_color=point_color)


if __name__ == "__main__":
    start_time = time.time()
    R_x, R_y, R_z = FuD.all_point(FuD.read_data()[0])
    # Create 3D surface
    surface = create_surface(R_x, R_y, R_z)

    #filework
    file_to_delete = 'PortsDATA_ver2.xlsx'
    input_file_path = os.path.join('Input_data', 'PreInput', 'PrePortsCoordinate.xlsx')
    output_file_path = 'PortsDATA_ver2.xlsx'

    #delete old file
    delete(file_to_delete)    
  
    #create current sheet
    output_sheet, output_workbook = create_new_file(input_file_path, output_file_path)
 
    #do task
    output_file_path = Module(output_sheet)






    Pname = name(output_sheet)


    # Получаем координаты из output_sheet
    X_1 = get_column_data(output_sheet, 5)
    Y_1 = get_column_data(output_sheet, 6)
    Z_1 = get_column_data(output_sheet, 7)
    X_2 = get_column_data(output_sheet, 8)
    Y_2 = get_column_data(output_sheet, 9)
    Z_2 = get_column_data(output_sheet, 10)

    # Создаем график
    plotter = pv.Plotter()
    plotter.add_mesh(surface, color='cyan', show_edges=True, opacity=0.2)

    # Добавляем линии на график
    for x1, y1, z1, x2, y2, z2 in zip(X_1, Y_1, Z_1, X_2, Y_2, Z_2):
        origin = np.array([x1, y1, z1])
        direction = np.array([x2, y2, z2])
        plotter.add_mesh(pv.Line(origin, direction), color='blue')


    add_labels(plotter, np.array([X_1[:36], Y_1[:36], Z_1[:36]]), Pname[:36])

    plotter.show()




    #savefile
    output_workbook.save(output_file_path)
    output_workbook.close()
 
 
